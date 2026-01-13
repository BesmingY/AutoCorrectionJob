import pandas as pd
from openpyxl import load_workbook

def normalize_key(student_id, student_name):
    """规范化学生标识符，确保一致性"""
    # 将可能的数字转换为整数再转为字符串，以消除浮点精度问题
    if pd.notna(student_id) and isinstance(student_id, float):
        student_id = str(int(student_id))
    else:
        student_id = str(student_id)
    
    if pd.notna(student_name):
        student_name = str(student_name)
    else:
        student_name = ""
    
    # 去除首尾空白
    student_id = student_id.strip()
    student_name = student_name.strip()
    
    return f"{student_id}_{student_name}"

def match_and_copy_fixed(file1_path, file2_path, output_path="实验final_结果.xlsx"):
    """
    根据实际文件结构修正的匹配代码
    """
    
    print("=== 开始处理 ===")
    
    # 读取文件1，使用header=5（第6行作为表头，第7行开始数据）
    df1 = pd.read_excel(file1_path, header=5)
    
    # 读取文件2，使用header=5
    df2 = pd.read_excel(file2_path, header=5)
    
    print("\n=== 文件1结构 ===")
    print("列名:", df1.columns.tolist())
    print("前3行数据:")
    print(df1.head(3))
    
    print("\n=== 文件2结构 ===")
    print("列名:", df2.columns.tolist())
    print("前3行数据:")
    print(df2.head(3))
    
    # 提取学号和姓名
    # 根据调试结果，学号在'Unnamed: 1'，姓名在'Unnamed: 2'
    df1['学号'] = df1['Unnamed: 1'].apply(lambda x: str(int(x)) if pd.notna(x) and isinstance(x, float) else str(x)).str.strip()
    df1['姓名'] = df1['Unnamed: 2'].astype(str).str.strip()
    
    df2['学号'] = df2['Unnamed: 1'].apply(lambda x: str(int(x)) if pd.notna(x) and isinstance(x, float) else str(x)).str.strip()
    df2['姓名'] = df2['Unnamed: 2'].astype(str).str.strip()
    
    # 创建唯一键
    df1['key'] = df1.apply(lambda row: f"{row['学号']}_{row['姓名']}", axis=1)
    df2['key'] = df2.apply(lambda row: f"{row['学号']}_{row['姓名']}", axis=1)
    
    print(f"\n=== 数据统计 ===")
    print(f"文件1记录数: {len(df1)}")
    print(f"文件2记录数: {len(df2)}")
    print(f"文件1前5个学生: {df1['key'].head().tolist()}")
    print(f"文件2前5个学生: {df2['key'].head().tolist()}")
    
    # 创建数据映射字典
    data_map = {}
    for idx, row in df1.iterrows():
        key = row['key']
        # 获取文件1的5-9次作业（对应列'5','6','7','8','9'）
        # 这些列名来自调试输出
        assignments = [
            row.get('5', None),
            row.get('6', None),
            row.get('7', None),
            row.get('8', None),
            row.get('9', None)
        ]
        data_map[key] = assignments
    
    print(f"\n数据映射创建完成，共 {len(data_map)} 条记录")
    
    # 查看部分映射数据
    print("\n文件1部分数据示例：")
    sample_keys = list(data_map.keys())[:3]
    for key in sample_keys:
        print(f"{key}: {data_map[key]}")
    
    # 使用openpyxl直接操作Excel文件2
    wb2 = load_workbook(file2_path)
    ws2 = wb2.active
    
    # 找到文件2中需要填充的列位置
    # 根据你的描述和调试结果，文件2的列结构不同
    # 我们需要找到列名为'5','6','7','8','9'的列索引
    print("\n=== 查找文件2的列位置 ===")
    
    # 先读取表头行（第6行，Excel行号6）
    header_row = 6  # Excel中第6行是表头（因为header=5）
    column_mapping = {}
    
    # 遍历所有列，找到目标列的位置
    for col in range(1, ws2.max_column + 1):
        cell_value = ws2.cell(row=header_row, column=col).value
        if cell_value in ['5', '6', '7', '8', '9']:
            column_mapping[cell_value] = col
            print(f"找到列 '{cell_value}' 在位置 {col} (Excel列 {chr(64+col)})")
    
    # 如果没找到，尝试其他方法
    if not column_mapping:
        print("未找到标准列名，尝试根据描述查找...")
        # 根据你的描述：文件2的D到I是1-6列，K-M是7,8,9列
        # 但实际结构可能不同，需要调整
        
        # 假设：第4列开始是作业成绩
        # 让我们打印列结构
        print("\n文件2列结构:")
        for col in range(1, min(20, ws2.max_column + 1)):
            cell_value = ws2.cell(row=header_row, column=col).value
            print(f"列 {col} ({chr(64+col)}): {cell_value}")
    
    # 简化：假设我们需要填充的列是H,I,K,L,M (8,9,11,12,13)
    # 对应作业5-9
    target_columns = [8, 9, 11, 12, 13]  # H,I,K,L,M
    print(f"\n目标列位置: {target_columns} (H,I,K,L,M)")
    
    # 开始匹配和复制
    not_found = []
    found_count = 0
    
    # 从第7行开始（数据行，Excel行号7）
    start_data_row = 7
    
    for excel_row in range(start_data_row, ws2.max_row + 1):
        # 获取学号和姓名
        student_id = ws2.cell(row=excel_row, column=2).value  # B列
        student_name = ws2.cell(row=excel_row, column=3).value  # C列
        
        if pd.isna(student_id) or pd.isna(student_name):
            continue

        # 使用专门的函数标准化键值，确保与data_map中的键一致
        key = normalize_key(student_id, student_name)
        
        if key in data_map:
            assignments = data_map[key]
            # 复制5-9次作业到目标列
            for i in range(5):
                target_col = target_columns[i]
                ws2.cell(row=excel_row, column=target_col).value = assignments[i]
            found_count += 1
            print(f"匹配成功: {key}")  # 添加调试信息
        else:
            not_found.append((student_id, student_name))
    
    # 保存结果
    wb2.save(output_path)
    
    print(f"\n=== 处理结果 ===")
    print(f"成功匹配: {found_count} 条记录")
    print(f"未匹配到: {len(not_found)} 条记录")
    
    if not_found:
        print("\n未找到的学号和姓名（前20个）:")
        for i, (student_id, student_name) in enumerate(not_found[:20]):
            print(f"{i+1}. 学号: {student_id}, 姓名: {student_name}")
        if len(not_found) > 20:
            print(f"... 还有 {len(not_found)-20} 条未显示")
    
    print(f"\n结果已保存到: {output_path}")
    
    return not_found

# 运行代码
if __name__ == "__main__":
    # 更新文件路径
    file1_path = r"D:\course\C++\实验_updated.xlsx"
    file2_path = r"D:\course\C++\实验final.xlsx"
    output_path = r"D:\course\C++\实验final_结果.xlsx"
    
    try:
        not_found = match_and_copy_fixed(file1_path, file2_path, output_path)
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()