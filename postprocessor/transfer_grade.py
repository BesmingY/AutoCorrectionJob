import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

def transfer_grades(excel_file, csv_file, output_file=None, unmatched_file=None):
    """
    将CSV文件中的成绩誊写到Excel文件中，并记录未匹配的记录
    
    参数:
    excel_file: Excel文件路径
    csv_file: CSV文件路径  
    output_file: 输出Excel文件路径（可选）
    unmatched_file: 未匹配记录输出文件路径（可选）
    """
    
    # 设置默认输出文件名
    if output_file is None:
        name, ext = os.path.splitext(excel_file)
        output_file = f"{name}_updated{ext}"
    
    if unmatched_file is None:
        name, ext = os.path.splitext(csv_file)
        unmatched_file = f"{name}_unmatched{ext}"
    
    try:
        # 读取CSV文件（指定UTF-8编码）
        print("正在读取CSV文件...")
        csv_df = pd.read_csv(csv_file, encoding='utf-8')
        print(f"成功读取CSV文件，共{len(csv_df)}条记录")
        
        # 创建一个副本来追踪每条CSV记录的匹配状态
        tracking_df = csv_df.copy()
        tracking_df['匹配状态'] = "未处理"  # 初始状态
        print(f"当前未匹配的记录数: {len(tracking_df)}")
        
        # 读取Excel文件
        print("正在读取Excel文件...")
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # 平时成绩所在列（M列，第13列）
        usual_grade_col = 24
        
        # 数据起始行
        start_row = 6
        
        # 更新计数器
        updated_count = 0
        
        print("开始匹配并更新成绩...")
        
        # 遍历Excel中的每一行学生数据
        for row in range(start_row, ws.max_row + 1):
            student_id = ws.cell(row=row, column=2).value  # 学号（B列）
            student_name = ws.cell(row=row, column=3).value  # 姓名（C列）

            # 跳过空行
            if student_id is None or student_name is None:
                continue

            # 查找CSV中匹配的记录
            match_mask = (
                (tracking_df['学号'].astype(str) == str(student_id)) &
                (tracking_df['姓名'].astype(str) == str(student_name))
            )

            if match_mask.any():
                # 获取成绩并更新Excel
                grade = tracking_df.loc[match_mask, '得分'].values[0]
                ws.cell(row=row, column=usual_grade_col).value = grade
                updated_count += 1

                # 标记该CSV记录为已匹配
                tracking_df.loc[match_mask, '匹配状态'] = "已匹配"

                print(f"已更新：{student_name} ({student_id}) - 成绩：{grade}")
            else:
                print(f"未找到匹配记录：{student_name} ({student_id})")

        # 保存更新后的Excel文件
        wb.save(output_file)

        # 处理CSV中有但Excel中没有的记录（即未被标记为“已匹配”的）
        unmatched_records = tracking_df[tracking_df['匹配状态'] == "未处理"]
        print(f"未匹配的记录数: {len(unmatched_records)}")
        print(f"匹配的记录数: {len(tracking_df[tracking_df['匹配状态'] == '已匹配'])}")
        print(f"总记录数: {len(tracking_df)}")
        if not unmatched_records.empty:
            unmatched_records = unmatched_records.copy()
            unmatched_records['未匹配原因'] = "Excel中无此学生"
            unmatched_records.to_csv(unmatched_file, index=False, encoding='utf-8-sig')
            print(f"\n保存了 {len(unmatched_records)} 条未匹配记录到: {unmatched_file}")
            print(f"文件编码: UTF-8 with BOM (utf-8-sig)")
        else:
            print("\n所有CSV记录都已成功匹配！")

        # 输出统计信息
        total_csv = len(csv_df)
        matched_count = len(tracking_df[tracking_df['匹配状态'] == "已匹配"])
        unmatched_count = len(unmatched_records)
        
        print(f"\n成绩更新完成！")
        print(f"CSV总记录数: {total_csv}")
        print(f"成功更新 {updated_count} 条记录")
        print(f"未匹配记录: {unmatched_count} 条记录")
        print(f"输出文件：{output_file}")

        # 数据完整性校验
        if matched_count + unmatched_count == total_csv:
            print("✓ 数据统计平衡：已匹配 + 未匹配 = 总记录数")
        else:
            print("⚠ 数据统计不平衡，请检查数据")

        return True

    except Exception as e:
        print(f"处理过程中出现错误：{str(e)}")
        return False

def check_duplicates(csv_file):
    """
    检查CSV文件中是否有重复的学号姓名组合
    """
    try:
        # 读取CSV文件
        df = pd.read_csv(csv_file, encoding='utf-8')
        print(f"成功读取CSV文件，共{len(df)}条记录")
        
        # 创建一个包含学号和姓名的列组合
        df['学号_姓名'] = df['学号'].astype(str) + '_' + df['姓名'].astype(str)
        
        # 查找重复的学号姓名组合
        duplicates = df[df.duplicated('学号_姓名', keep=False)]
        
        if not duplicates.empty:
            print(f"\n发现 {len(duplicates)} 条重复的学号姓名组合:")
            # 按学号姓名组合分组显示
            for combo in duplicates['学号_姓名'].unique():
                combo_records = duplicates[duplicates['学号_姓名'] == combo]
                if len(combo_records) > 1:
                    print(f"\n学号姓名组合 '{combo}' 出现 {len(combo_records)} 次:")
                    for idx, record in combo_records.iterrows():
                        print(f"  - 得分: {record['得分']}")
        else:
            print("\n未发现重复的学号姓名组合")
            
        # 显示所有唯一的学号姓名组合数量
        unique_combinations = df['学号_姓名'].nunique()
        print(f"\n共有 {unique_combinations} 个唯一的学号姓名组合")
        print(f"原始记录数: {len(df)}")
        
        return duplicates
        
    except Exception as e:
        print(f"处理过程中出现错误：{str(e)}")
        return None
def main():
    # 文件路径 - 请根据实际情况修改
    excel_file = r"D:\course\C++\实验.xlsx"  # 您的Excel文件
    csv_file = r"D:\course\C++\grading_results.csv"  # 您的CSV文件
    
    # 未匹配记录输出文件
    unmatched_file = r"D:\course\C++\自动批改作业agent\unmatched_records.csv"
    
    # 检查文件是否存在
    if not os.path.exists(excel_file):
        print(f"错误：Excel文件 '{excel_file}' 不存在")
        return
        
    if not os.path.exists(csv_file):
        print(f"错误：CSV文件 '{csv_file}' 不存在")
        return
    
    # 执行成绩转移
    success = transfer_grades(
        excel_file=excel_file,
        csv_file=csv_file,
        unmatched_file=unmatched_file
    )
    
    if success:
        print("\n操作完成！")
    else:
        print("\n操作失败，请检查文件格式和内容。")
        
    print("\n\n\n正在检查CSV文件中的重复记录...")
    check_duplicates(csv_file)

if __name__ == "__main__":
    main()